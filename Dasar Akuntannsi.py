import streamlit as st
import pandas as pd
from io import BytesIO
import uuid

# ================== KONFIGURASI ==================
st.set_page_config(page_title="Aplikasi Akuntansi", layout="wide")

# ================== SESSION STATE ==================
if "data" not in st.session_state:
    st.session_state.data = []

if "edit_id" not in st.session_state:
    st.session_state.edit_id = None

# ================== MASTER JENIS AKUN ==================
JENIS_AKUN = ["Aset", "Kewajiban", "Modal", "Pendapatan", "Beban"]

# ================== SIDEBAR MENU ==================
menu = st.sidebar.selectbox(
    "📂 Menu",
    [
        "🏠 Home",
        "📘 Jurnal Umum",
        "📑 Lihat Semua",
        "💾 Simpan ke Excel"
    ]
)

# ================== HOME ==================
if menu == "🏠 Home":
    st.title("📊 Aplikasi Akuntansi Sederhana")
    
# ================== JURNAL UMUM ==================
elif menu == "📘 Jurnal Umum":
    st.title("📘 Jurnal Umum")

    # ===== FORM INPUT =====
    with st.form("form_jurnal"):
        tanggal = st.date_input("Tanggal")

        col1, col2 = st.columns(2)
        with col1:
            akun_debit = st.text_input("Akun Debit")
            jenis_debit = st.selectbox("Jenis Akun Debit", JENIS_AKUN)
        with col2:
            akun_kredit = st.text_input("Akun Kredit")
            jenis_kredit = st.selectbox("Jenis Akun Kredit", JENIS_AKUN)

        jumlah = st.number_input("Jumlah", min_value=0.0)
        simpan = st.form_submit_button("Simpan Transaksi")

        if simpan:
            if akun_debit == "" or akun_kredit == "":
                st.error("Nama akun tidak boleh kosong")
            elif jumlah <= 0:
                st.error("Jumlah harus lebih dari 0")
            else:
                trx_id = str(uuid.uuid4())

                st.session_state.data.append({
                    "ID": trx_id,
                    "Tanggal": tanggal,
                    "Akun": akun_debit,
                    "Jenis Akun": jenis_debit,
                    "Debit": jumlah,
                    "Kredit": 0
                })
                st.session_state.data.append({
                    "ID": trx_id,
                    "Tanggal": tanggal,
                    "Akun": akun_kredit,
                    "Jenis Akun": jenis_kredit,
                    "Debit": 0,
                    "Kredit": jumlah
                })
                st.success("✅ Transaksi berhasil disimpan (SEIMBANG)")

    df = pd.DataFrame(st.session_state.data)

    # ===== TABEL JURNAL =====
    st.subheader("📋 Data Jurnal Umum")
    if df.empty:
        st.info("Belum ada transaksi")
    else:
        st.dataframe(df, use_container_width=True)

        # ===== KELOLA TRANSAKSI =====
        st.subheader("✏️ Kelola Transaksi")
        pilih_id = st.selectbox("Pilih ID Transaksi", df["ID"].unique())

        col1, col2 = st.columns(2)
        with col1:
            if st.button("🗑️ Hapus Transaksi"):
                st.session_state.data = [
                    d for d in st.session_state.data if d["ID"] != pilih_id
                ]
                st.experimental_rerun()

        with col2:
            if st.button("✏️ Edit Transaksi"):
                st.session_state.edit_id = pilih_id

        # ===== FORM EDIT =====
        if st.session_state.edit_id == pilih_id:
            trx = df[df["ID"] == pilih_id]
            debit_row = trx[trx["Debit"] > 0].iloc[0]
            kredit_row = trx[trx["Kredit"] > 0].iloc[0]

            st.markdown("### ✏️ Edit Transaksi")
            with st.form("form_edit"):
                tanggal_baru = st.date_input("Tanggal", debit_row["Tanggal"])

                col1, col2 = st.columns(2)
                with col1:
                    akun_debit_baru = st.text_input("Akun Debit", debit_row["Akun"])
                    jenis_debit_baru = st.selectbox(
                        "Jenis Akun Debit",
                        JENIS_AKUN,
                        index=JENIS_AKUN.index(debit_row["Jenis Akun"])
                    )
                with col2:
                    akun_kredit_baru = st.text_input("Akun Kredit", kredit_row["Akun"])
                    jenis_kredit_baru = st.selectbox(
                        "Jenis Akun Kredit",
                        JENIS_AKUN,
                        index=JENIS_AKUN.index(kredit_row["Jenis Akun"])
                    )

                jumlah_baru = st.number_input("Jumlah", value=float(debit_row["Debit"]))
                update = st.form_submit_button("🔄 Update")

                if update:
                    st.session_state.data = [
                        d for d in st.session_state.data if d["ID"] != pilih_id
                    ]
                    st.session_state.data.append({
                        "ID": pilih_id,
                        "Tanggal": tanggal_baru,
                        "Akun": akun_debit_baru,
                        "Jenis Akun": jenis_debit_baru,
                        "Debit": jumlah_baru,
                        "Kredit": 0
                    })
                    st.session_state.data.append({
                        "ID": pilih_id,
                        "Tanggal": tanggal_baru,
                        "Akun": akun_kredit_baru,
                        "Jenis Akun": jenis_kredit_baru,
                        "Debit": 0,
                        "Kredit": jumlah_baru
                    })
                    st.session_state.edit_id = None
                    st.experimental_rerun()

# ================== LIHAT SEMUA ==================
# ================== LIHAT SEMUA ==================
elif menu == "📑 Lihat Semua":
    st.title("📑 Semua Laporan")

    df = pd.DataFrame(st.session_state.data)

    if df.empty:
        st.warning("Belum ada data transaksi")
    else:
        # ================== JURNAL UMUM ==================
        st.subheader("📘 Jurnal Umum")
        st.dataframe(df, use_container_width=True)

        # ================== BUKU BESAR ==================
        st.subheader("📗 Buku Besar")
        buku_besar = (
            df.groupby(["Jenis Akun", "Akun"])[["Debit", "Kredit"]]
            .sum()
            .reset_index()
        )
        st.dataframe(buku_besar, use_container_width=True)

        # ================== LAPORAN LABA RUGI ==================
        st.subheader("📈 Laporan Laba Rugi")

        pendapatan = buku_besar[buku_besar["Jenis Akun"] == "Pendapatan"]
        beban = buku_besar[buku_besar["Jenis Akun"] == "Beban"]

        total_pendapatan = pendapatan["Kredit"].sum() - pendapatan["Debit"].sum()
        total_beban = beban["Debit"].sum() - beban["Kredit"].sum()
        laba_bersih = total_pendapatan - total_beban

        col1, col2, col3 = st.columns(3)
        col1.metric("Total Pendapatan", f"Rp {total_pendapatan:,.0f}")
        col2.metric("Total Beban", f"Rp {total_beban:,.0f}")
        col3.metric("Laba Bersih", f"Rp {laba_bersih:,.0f}")

        # ================== NERACA ==================
        st.subheader("⚖️ Neraca")

        aset = buku_besar[buku_besar["Jenis Akun"] == "Aset"]
        kewajiban = buku_besar[buku_besar["Jenis Akun"] == "Kewajiban"]
        modal = buku_besar[buku_besar["Jenis Akun"] == "Modal"]

        total_aset = aset["Debit"].sum() - aset["Kredit"].sum()
        total_kewajiban = kewajiban["Kredit"].sum() - kewajiban["Debit"].sum()
        total_modal = modal["Kredit"].sum() - modal["Debit"].sum()

        # Modal ditambah Laba Bersih
        modal_akhir = total_modal + laba_bersih
        total_pasiva = total_kewajiban + modal_akhir

        col1, col2 = st.columns(2)
        col1.metric("Total Aset", f"Rp {total_aset:,.0f}")
        col2.metric("Total Kewajiban + Modal", f"Rp {total_pasiva:,.0f}")

        if total_aset == total_pasiva:
            st.success("✔ Neraca Seimbang")
        else:
            st.error("❌ Neraca Tidak Seimbang")

# ================== SIMPAN EXCEL ==================
# ================== SIMPAN EXCEL ==================
elif menu == "💾 Simpan ke Excel":
    st.title("💾 Simpan ke Excel")

    df = pd.DataFrame(st.session_state.data)

    if df.empty:
        st.warning("Belum ada data")
    else:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # ================== BUKU BESAR ==================
        buku_besar = (
            df.groupby(["Jenis Akun", "Akun"])[["Debit", "Kredit"]]
            .sum()
            .reset_index()
        )

        # ================== LABA RUGI ==================
        pendapatan = buku_besar[buku_besar["Jenis Akun"] == "Pendapatan"]
        beban = buku_besar[buku_besar["Jenis Akun"] == "Beban"]

        total_pendapatan = pendapatan["Kredit"].sum() - pendapatan["Debit"].sum()
        total_beban = beban["Debit"].sum() - beban["Kredit"].sum()
        laba_bersih = total_pendapatan - total_beban

        laba_rugi_df = pd.DataFrame({
            "Keterangan": ["Total Pendapatan", "Total Beban", "Laba Bersih"],
            "Jumlah": [total_pendapatan, total_beban, laba_bersih]
        })

        # ================== NERACA ==================
        aset = buku_besar[buku_besar["Jenis Akun"] == "Aset"]
        kewajiban = buku_besar[buku_besar["Jenis Akun"] == "Kewajiban"]
        modal = buku_besar[buku_besar["Jenis Akun"] == "Modal"]

        total_aset = aset["Debit"].sum() - aset["Kredit"].sum()
        total_kewajiban = kewajiban["Kredit"].sum() - kewajiban["Debit"].sum()
        total_modal = modal["Kredit"].sum() - modal["Debit"].sum()

        modal_akhir = total_modal + laba_bersih
        total_pasiva = total_kewajiban + modal_akhir

        neraca_df = pd.DataFrame({
            "Keterangan": [
                "Total Aset",
                "Total Kewajiban",
                "Modal Akhir",
                "Total Kewajiban + Modal"
            ],
            "Jumlah": [
                total_aset,
                total_kewajiban,
                modal_akhir,
                total_pasiva
            ]
        })

        # ================== EXPORT & FORMAT ==================
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Jurnal Umum", index=False, startrow=2)
            buku_besar.to_excel(writer, sheet_name="Buku Besar", index=False, startrow=2)
            laba_rugi_df.to_excel(writer, sheet_name="Laba Rugi", index=False, startrow=2)
            neraca_df.to_excel(writer, sheet_name="Neraca", index=False, startrow=2)

            workbook = writer.book

            def format_sheet(sheet_name, title):
                ws = workbook[sheet_name]

                # Judul
                ws.merge_cells("A1:D1")
                ws["A1"] = title
                ws["A1"].font = Font(bold=True, size=14)
                ws["A1"].alignment = Alignment(horizontal="center")

                header_fill = PatternFill("solid", fgColor="BDD7EE")
                border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )

                for cell in ws[3]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.border = border

                for row in ws.iter_rows(min_row=4):
                    for cell in row:
                        cell.border = border
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = 'Rp #,##0'

                # Auto width
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_length + 2

            format_sheet("Jurnal Umum", "LAPORAN JURNAL UMUM")
            format_sheet("Buku Besar", "LAPORAN BUKU BESAR")
            format_sheet("Laba Rugi", "LAPORAN LABA RUGI")
            format_sheet("Neraca", "LAPORAN NERACA")

        output.seek(0)

        st.download_button(
            "📥 Download Laporan Excel (Format Rapi)",
            data=output,
            file_name="laporan_akuntansi_format_rapi.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
